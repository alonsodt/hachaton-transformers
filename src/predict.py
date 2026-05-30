"""Genera el envio: reentrena con TODO el historico y rellena la plantilla .xlsx.

    python -m src.predict                      # usa el mejor modelo (drift_full)
    python -m src.predict --model blend_full_0.7

CLAVE (formato de fecha): la columna 'Date' del template viene como TEXTO
'YYYY-MM-DD' (formato '@', sin hora). Abrimos el .xlsx con openpyxl y escribimos
SOLO las celdas pred_index_* , dejando la columna Date intacta byte a byte. Asi la
clave de fecha casa exacta con el corrector y nunca aparece '00:00:00'. (El bug
anterior reconvertia Date a datetime con pandas y la reescribia.)
"""
from __future__ import annotations

import argparse
from datetime import datetime

import numpy as np
import openpyxl

from . import config as C
from . import data as D
from . import models as M


def build_predictions(model: str) -> dict[str, np.ndarray]:
    """Devuelve {pred_col: array} alineado con las filas de la hoja 'submission'."""
    idx = D.load_indices()

    wb = openpyxl.load_workbook(C.TEMPLATE_XLSX)
    ws = wb[C.TEMPLATE_SHEET]
    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    date_col_i = header.index(C.DATE_COL) + 1
    # fechas del template (texto) -> Timestamps solo para proyectar el drift
    import pandas as pd
    raw_dates = [ws.cell(r, date_col_i).value for r in range(2, ws.max_row + 1)]
    future_dates = pd.DatetimeIndex(pd.to_datetime(raw_dates))

    fn = M.REGISTRY[model]
    preds = {}
    for t in C.TARGETS:
        series = idx[t].dropna()
        preds[C.SUB_COL[t]] = np.asarray(fn(series, future_dates), dtype=float)
    return preds


def write_submission(model: str, out_path) -> int:
    """Escribe el envio preservando la columna Date; devuelve nº de filas."""
    preds = build_predictions(model)

    wb = openpyxl.load_workbook(C.TEMPLATE_XLSX)
    ws = wb[C.TEMPLATE_SHEET]
    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col_idx = {name: i + 1 for i, name in enumerate(header)}
    n = ws.max_row - 1  # filas de datos

    for pred_col, values in preds.items():
        assert len(values) == n, f"{pred_col}: {len(values)} vs {n} filas"
        assert np.isfinite(values).all(), f"NaN/inf en {pred_col}"
        assert (values > 0).all(), f"valores no positivos en {pred_col}"
        ci = col_idx[pred_col]
        for r, v in enumerate(values, start=2):
            ws.cell(r, ci).value = float(v)

    wb.save(out_path)
    return n


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--model", default="drift_full", choices=list(M.REGISTRY))
    ap.add_argument("--out", default=None)
    args = ap.parse_args()

    C.SUBMISSIONS.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d")
    out = args.out or (C.SUBMISSIONS / f"submission_{stamp}_{args.model}.xlsx")
    n = write_submission(args.model, out)

    # re-leer y enseñar cabecera/cola para verificar visualmente
    wb = openpyxl.load_workbook(out)
    ws = wb[C.TEMPLATE_SHEET]
    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    print(f"[predict] modelo={args.model} | filas={n}")
    print(f"[predict] envio escrito en: {out}")
    print("  " + " | ".join(str(h) for h in header))
    for r in list(range(2, 5)) + [ws.max_row]:
        cells = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        fmt = [cells[0]] + [f"{v:,.0f}" if isinstance(v, float) else v for v in cells[1:]]
        print("  " + " | ".join(str(x) for x in fmt))


if __name__ == "__main__":
    main()
